#!/usr/bin/env python3
"""Generator nowoczesnego dashboardu HTML z pliku audytu xlsx.
Użycie: python3 scripts/dashboard.py outputs/AUDYT_X_v1.xlsx [outdir]
Tworzy samodzielny plik HTML (Tailwind + Chart.js z CDN) — otwierasz w przeglądarce.
"""
import sys, os, json, re, html
from openpyxl import load_workbook

def num(v):
    return v if isinstance(v,(int,float)) else None

# Czyści nagłówek streszczenia z prefiksów "STRESZCZENIE ZARZADCZE" / "AUDYT PRZETARGOWY [v2]"
# i emoji — zostaje realna nazwa inwestycji do tabel.
_EMOJI=re.compile(r"[\U0001F000-\U0001FAFF☀-➿️]")
def clean_title(s):
    s=_EMOJI.sub(" ", str(s or "")).strip()
    prev=None
    while s!=prev:
        prev=s
        s=re.sub(r"(?i)^(streszczenie\s+zarz[aą]dcze|audyt\s+przetargow[ay](\s+v?\d+)?)\b","",s)
        s=s.lstrip(" \t-—–:•").strip()
    return re.sub(r"\s{2,}"," ",s) or "Audyt przetargowy"

# "1,85" / "1 850" -> 1850000 (z banera "X,XX mln PLN netto")
def _mln(s):
    try: return int(round(float(str(s).replace(" ","").replace(",","."))*1e6))
    except Exception: return None

def find_sheet(wb, *frags):
    for s in wb.sheetnames:
        for f in frags:
            if f.lower() in s.lower(): return wb[s]
    return None

def cells(ws):
    out=[]
    for row in ws.iter_rows():
        out.append([c.value for c in row])
    return out

def parse(path):
    wb=load_workbook(path, data_only=True)
    d={"file":os.path.basename(path), "groups":[], "materials":[], "risks":[], "questions":[],
       "meta":{}, "value":{}, "mat_total":None, "status_counts":{}}
    # --- nazwa / wartość / meta z arkusza streszczenia (układ klucz-wartość B/C) ---
    s01=find_sheet(wb,"01_STRESZCZENIE","STRESZCZENIE")
    if s01:
        rows01=cells(s01)
        a1=str((rows01[0][0] if rows01 and rows01[0] else "") or "")
        a2=str((rows01[1][0] if len(rows01)>1 and rows01[1] else "") or "")
        d["meta"]["project"]=clean_title(a1)
        d["value"]["banner"]=a2
        kv={}
        for r in rows01:
            if len(r)>=3 and r[1] and r[2]:
                lab=str(r[1]).strip().lower(); val=str(r[2]).strip()
                if lab and val and lab not in kv: kv[lab]=val
        def kvget(*frags):
            for lab,val in kv.items():
                if any(f in lab for f in frags): return val
            return ""
        flat=" ".join(str(c) for row in rows01 for c in row if c)
        # termin składania ofert (dd.mm.yyyy lub yyyy-mm-dd, opcjonalnie godz.)
        traw=kvget("termin skład","termin sklad","skladani ofert","składani ofert","termin oferty","termin ofert")
        _DATE=r"(?:(?:0?[1-9]|[12]\d|3[01])[.\-/](?:0?[1-9]|1[0-2])[.\-/]20\d{2}|20\d{2}-(?:0[1-9]|1[0-2])-(?:0[1-9]|[12]\d|3[01]))"
        mt=re.search(r"("+_DATE+r")(?:[^\d]{0,8}(godz\.?\s*\d{1,2}[:.]\d{2}))?",traw)
        if mt: d["meta"]["termin"]=(mt.group(1)+(" "+re.sub(r"\s+"," ",mt.group(2)) if mt.group(2) else "")).strip()
        # wadium
        wraw=kvget("wadium")
        if wraw:
            mw=re.search(r"(\d[\d\s]{2,}\s*(?:PLN|zł|zl))",wraw)
            d["meta"]["wadium"]=(re.sub(r"\s+"," ",mw.group(1)).strip() if mw else ("brak" if "BRAK" in wraw.upper() else wraw[:40]))
        # numer postępowania / BZP
        mbzp=re.search(r"(\d{4}/BZP\s*\d+|BZP\s*\d{6,})",flat)
        if mbzp: d["meta"]["bzp"]=re.sub(r"\s+"," ",mbzp.group(1)).strip()
        else:
            nr=kvget("numer post","nr post","znak spraw","nr spraw")
            if nr: d["meta"]["bzp"]=re.split(r"[(;|]",nr)[0].strip()[:32]
        # rekomendacja / werdykt
        rek=kvget("werdykt","rekomendacja")
        if not rek:
            mr=re.search(r"((?:WARUNKOWO\s+)?(?:NIE\s+)?SK[ŁL]ADA[ĆC][^.\n]{0,90})",flat,re.I)
            if mr: rek=mr.group(1).strip()
        if rek: d["meta"]["rekomendacja"]=rek
        zam=kvget("zamawiaj")
        if zam: d["meta"]["zamawiajacy"]=zam[:80]
        # wartość PRIMARY z banera A2 ("X,XX – Y,YY mln PLN netto (... mln brutto)")
        mb=re.search(r"([\d,]+)\s*[–\-]\s*([\d,]+)\s*mln\s*PLN\s*netto",a2)
        if mb:
            d["value"]["netto_od"]=_mln(mb.group(1)); d["value"]["netto_do"]=_mln(mb.group(2))
        mbb=re.search(r"\(\s*([\d,]+)\s*[–\-]\s*([\d,]+)\s*mln\s*brutto",a2)
        if mbb:
            d["value"]["brutto_od"]=_mln(mbb.group(1)); d["value"]["brutto_do"]=_mln(mbb.group(2))
        if d["value"].get("netto_od") is None:
            mfn=re.search(r"([\d.,]+)\s*[–\-]\s*([\d.,]+)\s*mln\s*PLN\s*netto",flat)
            if mfn:
                d["value"]["netto_od"]=_mln(mfn.group(1)); d["value"]["netto_do"]=_mln(mfn.group(2))
    # --- 25 kalkulacja: grupy do wykresu + RAZEM/SUMA jako FALLBACK wartości ---
    s25=find_sheet(wb,"25_KALKULACJA","KALKULACJA")
    if s25:
        rows=cells(s25); hdr=None
        for i,r in enumerate(rows):
            if r and any(str(x).strip() in ("Grupa kosztowa","Branża","Branza","Element") for x in r if x): hdr=i; break
        if hdr is not None:
            for r in rows[hdr+1:]:
                b=str(r[1]) if len(r)>1 and r[1] else ""
                if not b: continue
                bu=b.upper()
                if "RAZEM" in bu or "SUMA" in bu:
                    if d["value"].get("netto_od") is None:
                        nums=[num(x) for x in r if num(x) is not None and num(x)>=1000]
                        if nums:
                            d["value"]["netto_od"]=nums[0]
                            d["value"]["netto_do"]=nums[1] if len(nums)>1 else nums[0]
                    break
                od=num(r[2]) if len(r)>2 else None
                do=num(r[3]) if len(r)>3 else None
                if od is not None or do is not None:
                    d["groups"].append({"name":b,"od":od or do or 0,"do":do or od or 0})
        if d["value"].get("netto_od") is None:
            for r in rows:
                lab=" ".join(str(x) for x in r[:3] if x).upper()
                if ("RAZEM" in lab or "SUMA" in lab) and "NETTO" in lab:
                    nums=[num(x) for x in r if num(x) is not None and num(x)>=1000]
                    if nums:
                        d["value"]["netto_od"]=nums[0]
                        d["value"]["netto_do"]=nums[1] if len(nums)>1 else nums[0]
                        break
    # --- M00 materiały ---
    m00=find_sheet(wb,"M00_ZESTAWIENIE","ZESTAWIENIE_ZBIORCZE")
    if m00:
        rows=cells(m00); hdr=None
        for i,r in enumerate(rows):
            if r and any(str(x).strip().lower() in ("materiał","material") for x in r if x): hdr=i; break
        if hdr is None: hdr=2
        # mapowanie kolumn po nagłówkach
        head=[str(x).strip().lower() if x else "" for x in rows[hdr]]
        def col(*names,default=None):
            for n in names:
                for j,h in enumerate(head):
                    if n in h: return j
            return default
        ci={"mat":col("materiał","material",default=2),"branza":col("branża","branza",default=1),
            "jedn":col("jedn",default=4),"il":col("ilość","ilosc",default=5),
            "cena":col("cena",default=7),"wart":col("wartość","wartosc",default=8),"stat":col("status",default=9)}
        for r in rows[hdr+1:]:
            mat=r[ci["mat"]] if len(r)>ci["mat"] else None
            if not mat or "RAZEM" in str(mat).upper(): continue
            if str(r[1] if len(r)>1 else "").upper().startswith("RAZEM"): continue
            st=str(r[ci["stat"]]).split()[0] if len(r)>ci["stat"] and r[ci["stat"]] else "OK"
            d["materials"].append({
                "mat":str(mat),"branza":str(r[ci["branza"]] or "") if len(r)>ci["branza"] else "",
                "jedn":str(r[ci["jedn"]] or "") if len(r)>ci["jedn"] else "",
                "il":r[ci["il"]] if len(r)>ci["il"] else "",
                "cena":num(r[ci["cena"]]) if len(r)>ci["cena"] else None,
                "wart":num(r[ci["wart"]]) if len(r)>ci["wart"] else None,
                "stat":st})
            d["status_counts"][st]=d["status_counts"].get(st,0)+1
        d["mat_total"]=round(sum(m["wart"] or 0 for m in d["materials"]),2)
    # --- ryzyka (24 lub z 01) ---
    s24=find_sheet(wb,"24_RYZYKA","RYZYKA")
    src=s24 or s01
    if src:
        for r in cells(src):
            line=[str(x) for x in r if x]
            sev=None
            for cell in line:
                for k in ("KRYTYCZNA","WYSOKA","ŚREDNIA","SREDNIA","NISKA"):
                    if cell.strip().upper()==k: sev=k.replace("Ś","S"); break
            if sev:
                title=next((x for x in line if len(x)>12 and x.strip().upper()!=sev),"")
                desc=next((x for x in line if len(x)>40),"")
                if title: d["risks"].append({"sev":sev,"title":title[:120],"desc":desc[:240]})
    # dedup ryzyk
    seen=set(); d["risks"]=[r for r in d["risks"] if not (r["title"] in seen or seen.add(r["title"]))][:8]
    # --- pytania (23) ---
    s23=find_sheet(wb,"23_PYTANIA","PYTANIA")
    if s23:
        for r in cells(s23):
            for x in r:
                if x and isinstance(x,str) and (x.strip().startswith(("Prosimy","Czy","Wnosimy","P-")) and len(x)>25):
                    d["questions"].append(x.strip()[:260])
        d["questions"]=d["questions"][:6]
    return d

PLN=lambda v: ("{:,.0f}".format(v).replace(","," ")+" zł") if isinstance(v,(int,float)) else "—"
MLN=lambda v: ("{:.2f}".format(v/1e6).replace("."," ,".strip())) if isinstance(v,(int,float)) else "—"
def mln(v): return ("%.2f"%(v/1e6)).replace(".",",") if isinstance(v,(int,float)) else "—"

SEV_COLOR={"KRYTYCZNA":"#dc2626","WYSOKA":"#ea580c","SREDNIA":"#d97706","NISKA":"#65a30d"}
ST_COLOR={"OK":"#16a34a","DOPYTAC":"#d97706","DOPYTAĆ":"#d97706","KRYTYCZNA":"#dc2626",
          "BRAK_W_PRZEDMIARZE":"#2563eb","ROZBIEZNOSC":"#db2777","ROZBIEŻNOŚĆ":"#db2777","SREDNIA":"#d97706"}

def esc(s): return html.escape(str(s)) if s is not None else ""

def render(d):
    proj=esc(d["meta"].get("project","Audyt przetargowy"))
    rek=d["meta"].get("rekomendacja","")
    rek_bid = rek.upper().startswith("SK") and not rek.upper().startswith("NIE")
    nod,ndo=d["value"].get("netto_od"),d["value"].get("netto_do")
    bod,bdo=d["value"].get("brutto_od"),d["value"].get("brutto_do")
    val_net = (mln(nod)+" – "+mln(ndo)+" mln") if nod else "—"
    val_bru = (mln(bod)+" – "+mln(bdo)+" mln") if bod else "—"
    mats=d["materials"]; risks=d["risks"]; groups=d["groups"]; qs=d["questions"]
    krit=sum(1 for r in risks if r["sev"]=="KRYTYCZNA")
    # dane wykresów
    g_labels=json.dumps([g["name"][:34] for g in groups], ensure_ascii=False)
    g_od=json.dumps([round(g["od"]/1000) for g in groups]); g_do=json.dumps([round(g["do"]/1000) for g in groups])
    st_labels=json.dumps(list(d["status_counts"].keys()), ensure_ascii=False)
    st_vals=json.dumps(list(d["status_counts"].values()))
    st_colors=json.dumps([ST_COLOR.get(k,"#64748b") for k in d["status_counts"].keys()])
    # wiersze materiałów
    mat_rows="".join(
        f'<tr class="border-b border-slate-100 hover:bg-slate-50" data-s="{esc(m["stat"])}">'
        f'<td class="py-2 px-3 text-slate-700">{esc(m["mat"])}</td>'
        f'<td class="py-2 px-3 text-slate-500 text-xs">{esc(m["branza"])}</td>'
        f'<td class="py-2 px-3 text-right">{esc(m["il"])}</td>'
        f'<td class="py-2 px-3 text-slate-500">{esc(m["jedn"])}</td>'
        f'<td class="py-2 px-3 text-right">{PLN(m["cena"]) if m["cena"] else "—"}</td>'
        f'<td class="py-2 px-3 text-right font-semibold">{PLN(m["wart"]) if m["wart"] else "—"}</td>'
        f'<td class="py-2 px-3"><span class="text-xs font-bold px-2 py-0.5 rounded" style="background:{ST_COLOR.get(m["stat"],"#64748b")}22;color:{ST_COLOR.get(m["stat"],"#64748b")}">{esc(m["stat"])}</span></td></tr>'
        for m in mats)
    risk_cards="".join(
        f'<div class="rounded-xl border-l-4 bg-white p-4 shadow-sm" style="border-color:{SEV_COLOR.get(r["sev"],"#64748b")}">'
        f'<div class="flex items-center gap-2 mb-1"><span class="text-xs font-extrabold px-2 py-0.5 rounded text-white" style="background:{SEV_COLOR.get(r["sev"],"#64748b")}">{esc(r["sev"])}</span></div>'
        f'<p class="font-semibold text-slate-800 text-sm">{esc(r["title"])}</p>'
        f'<p class="text-slate-500 text-xs mt-1">{esc(r["desc"])}</p></div>'
        for r in risks)
    q_rows="".join(f'<li class="flex gap-3 py-2 border-b border-slate-100"><span class="text-orange-500 font-bold">?</span><span class="text-sm text-slate-600">{esc(q)}</span></li>' for q in qs)
    status_chips="".join(f'<button onclick="flt(this,\'{esc(k)}\')" class="chip text-xs font-bold px-3 py-1 rounded-full border" style="color:{ST_COLOR.get(k,"#64748b")};border-color:{ST_COLOR.get(k,"#64748b")}55">{esc(k)} · {v}</button>' for k,v in d["status_counts"].items())
    rek_txt = esc(rek) if rek else ("Rekomendacja w zakładce 01_STRESZCZENIE")

    return f"""<!doctype html><html lang="pl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{proj} — Dashboard audytu</title>
<script src="https://cdn.tailwindcss.com"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>body{{font-family:ui-sans-serif,system-ui,'Segoe UI',Roboto,sans-serif}}
.kpi{{background:linear-gradient(135deg,#0d2b4e,#13386a)}}</style></head>
<body class="bg-slate-100 text-slate-800">
<header class="bg-[#0d2b4e] text-white px-6 py-4 sticky top-0 z-20 shadow-lg">
 <div class="max-w-7xl mx-auto flex items-center justify-between gap-4 flex-wrap">
  <div><div class="text-xs uppercase tracking-widest text-orange-400 font-bold">Analizator dokumentacji przetargowej</div>
   <h1 class="text-lg md:text-xl font-extrabold leading-tight">{proj}</h1>
   <div class="text-xs text-slate-300 mt-0.5">{esc(d['meta'].get('bzp',''))} {('· '+esc(d['meta'].get('file',''))) if d['meta'].get('file') else ''}</div></div>
  <div class="px-4 py-2 rounded-xl font-extrabold text-sm {'bg-green-500' if rek_bid else 'bg-orange-500'} text-white whitespace-nowrap">{'✓ SKŁADAĆ' if rek_bid else '⚑ DECYZJA'}</div>
 </div></header>

<main class="max-w-7xl mx-auto px-4 md:px-6 py-6 space-y-6">
 <!-- KPI -->
 <section class="grid grid-cols-2 md:grid-cols-4 gap-4">
  <div class="kpi text-white rounded-2xl p-5 shadow"><div class="text-xs uppercase tracking-wider text-orange-300 font-bold">Wartość szacunkowa</div>
   <div class="text-2xl font-extrabold mt-1">{val_net}</div><div class="text-xs text-slate-300">netto · brutto {val_bru}</div></div>
  <div class="bg-white rounded-2xl p-5 shadow"><div class="text-xs uppercase tracking-wider text-slate-400 font-bold">Materiały (M00)</div>
   <div class="text-2xl font-extrabold mt-1 text-[#0d2b4e]">{PLN(d['mat_total'])}</div><div class="text-xs text-slate-400">{len(mats)} pozycji</div></div>
  <div class="bg-white rounded-2xl p-5 shadow"><div class="text-xs uppercase tracking-wider text-slate-400 font-bold">Ryzyka</div>
   <div class="text-2xl font-extrabold mt-1 text-red-600">{len(risks)}</div><div class="text-xs text-slate-400">{krit} krytyczne</div></div>
  <div class="bg-white rounded-2xl p-5 shadow"><div class="text-xs uppercase tracking-wider text-slate-400 font-bold">Termin / wadium</div>
   <div class="text-lg font-extrabold mt-1 text-[#0d2b4e]">{esc(d['meta'].get('termin','—'))}</div><div class="text-xs text-slate-400">wadium {esc(d['meta'].get('wadium','—'))}</div></div>
 </section>

 <!-- Rekomendacja -->
 <section class="rounded-2xl p-5 shadow bg-white border-l-8 {'border-green-500' if rek_bid else 'border-orange-500'}">
  <div class="text-xs uppercase tracking-wider text-slate-400 font-bold mb-1">Rekomendacja</div>
  <p class="text-slate-700">{rek_txt}</p></section>

 <div class="grid lg:grid-cols-3 gap-6">
  <!-- Wykres kosztów -->
  <section class="lg:col-span-2 bg-white rounded-2xl p-5 shadow">
   <h2 class="font-bold text-[#0d2b4e] mb-3">Struktura kosztów — widełki (tys. zł netto)</h2>
   <canvas id="cost" height="120"></canvas></section>
  <!-- Statusy materiałów -->
  <section class="bg-white rounded-2xl p-5 shadow">
   <h2 class="font-bold text-[#0d2b4e] mb-3">Statusy pozycji</h2>
   <canvas id="stat" height="160"></canvas></section>
 </div>

 <!-- Ryzyka -->
 <section><h2 class="font-bold text-[#0d2b4e] mb-3 text-lg">Najważniejsze ryzyka</h2>
  <div class="grid md:grid-cols-2 gap-3">{risk_cards or '<p class="text-slate-400">Brak danych o ryzykach.</p>'}</div></section>

 <!-- Pytania -->
 {('<section class="bg-white rounded-2xl p-5 shadow"><h2 class="font-bold text-[#0d2b4e] mb-2">Pytania do zamawiającego (gotowe)</h2><ul>'+q_rows+'</ul></section>') if qs else ''}

 <!-- Materiały -->
 <section class="bg-white rounded-2xl p-5 shadow">
  <div class="flex items-center justify-between gap-3 flex-wrap mb-3">
   <h2 class="font-bold text-[#0d2b4e] text-lg">Zestawienie materiałów (M00)</h2>
   <input id="q" oninput="srch()" placeholder="szukaj materiału…" class="border rounded-lg px-3 py-1.5 text-sm w-56">
  </div>
  <div class="flex gap-2 flex-wrap mb-3"><button onclick="flt(this,'')" class="chip text-xs font-bold px-3 py-1 rounded-full border border-slate-300 bg-slate-100">Wszystkie</button>{status_chips}</div>
  <div class="overflow-x-auto"><table class="w-full text-sm"><thead><tr class="text-left text-xs uppercase text-slate-400 border-b-2 border-slate-200">
   <th class="py-2 px-3">Materiał</th><th class="py-2 px-3">Branża</th><th class="py-2 px-3 text-right">Ilość</th><th class="py-2 px-3">Jedn.</th><th class="py-2 px-3 text-right">Cena</th><th class="py-2 px-3 text-right">Wartość</th><th class="py-2 px-3">Status</th></tr></thead>
   <tbody id="mtb">{mat_rows}</tbody></table></div></section>

 <footer class="text-center text-xs text-slate-400 py-6">Dashboard wygenerowany automatycznie z pliku audytu · Analizator dokumentacji przetargowej WSK</footer>
</main>
<script>
const gl={g_labels}, god={g_od}, gdo={g_do};
new Chart(document.getElementById('cost'),{{type:'bar',data:{{labels:gl,datasets:[
 {{label:'od',data:god,backgroundColor:'#1a4d80'}},{{label:'do',data:gdo,backgroundColor:'#d35400'}}]}},
 options:{{indexAxis:'y',plugins:{{legend:{{display:true}}}},scales:{{x:{{ticks:{{callback:v=>v+'k'}}}}}}}}}});
new Chart(document.getElementById('stat'),{{type:'doughnut',data:{{labels:{st_labels},datasets:[{{data:{st_vals},backgroundColor:{st_colors}}}]}},options:{{plugins:{{legend:{{position:'bottom'}}}}}}}});
function srch(){{let q=document.getElementById('q').value.toLowerCase();document.querySelectorAll('#mtb tr').forEach(r=>{{r.style.display=r.innerText.toLowerCase().includes(q)?'':'none'}})}}
function flt(b,s){{document.querySelectorAll('.chip').forEach(c=>c.classList.remove('ring-2'));b.classList.add('ring-2');document.querySelectorAll('#mtb tr').forEach(r=>{{r.style.display=(!s||r.dataset.s===s)?'':'none'}})}}
</script></body></html>"""

if __name__=="__main__":
    if len(sys.argv)<2: print("Uzycie: dashboard.py <plik.xlsx> [outdir]"); sys.exit(1)
    path=sys.argv[1]; outdir=sys.argv[2] if len(sys.argv)>2 else "/home/wskpawelw/audyt/dashboards"
    os.makedirs(outdir,exist_ok=True)
    d=parse(path)
    name=os.path.splitext(os.path.basename(path))[0]+".html"
    outp=os.path.join(outdir,name)
    open(outp,"w",encoding="utf-8").write(render(d))
    print("OK ->",outp)
    print(f"  materiałów: {len(d['materials'])} | ryzyk: {len(d['risks'])} | grup kosztów: {len(d['groups'])} | wartość: {d['value']}")


# ============ COVERAGE: co i jak zostało przeanalizowane (zakładka "Co przeanalizowano") ============
def _sheet_rows(wb, *frags):
    ws=find_sheet(wb, *frags)
    if not ws: return []
    return [[c.value for c in r] for r in ws.iter_rows()]

def _hdr_idx(rows, *keys):
    # naglowek = krotkie etykiety (<30 zn.); chroni przed zlapaniem banera-tytulu
    keys=[k.lower() for k in keys]
    for i,r in enumerate(rows[:8]):
        cells=[str(x).strip().lower() if x else "" for x in r]
        if any(any(k in c and len(c)<30 for c in cells) for k in keys): return i
    return None

def _col(head, *names, default=None):
    for n in names:
        for j,h in enumerate(head):
            if n in h: return j
    return default

def coverage(path):
    """Prowieniencja audytu: jakie dokumenty, czym przetworzone (OCR vision/tekst/przeczytane),
    co porównane (cross-ref), rozbieżności, braki, eksperci. Czyta arkusze 04/X01/X02/X03/07/08/E*."""
    from openpyxl import load_workbook
    wb=load_workbook(path, data_only=True)
    cov={"documents":[],"doc_stats":{},"ocr_rysunki":{},"ocr_decyzje":{},
         "crossref":{},"rozbieznosci":[],"braki":[],"experts":[],"model":"Opus 4.8 (vision)"}
    def cell(row,i,n=120):
        return str(row[i])[:n] if i is not None and len(row)>i and row[i] not in (None,"") else ""
    # 04 — inwentaryzacja dokumentów
    rows=_sheet_rows(wb,"04_INWENTARYZACJA","INWENTARYZACJA")
    hi=_hdr_idx(rows,"plik") if rows else None
    if hi is not None:
        head=[str(x).strip().lower() if x else "" for x in rows[hi]]
        cp=_col(head,"plik"); ct=_col(head,"typ"); cf=_col(head,"format"); cfo=_col(head,"folder"); cs=_col(head,"status")
        v=t=r=0
        for row in rows[hi+1:]:
            plik=cell(row,cp,120)
            if not plik or plik.strip() in ("","#"): continue
            st=cell(row,cs,40); low=st.lower()
            kind="vision" if "vision" in low else ("ocr" if "ocr" in low else "read")
            if kind=="vision": v+=1
            elif kind=="ocr": t+=1
            else: r+=1
            cov["documents"].append({"plik":plik,"typ":cell(row,ct,90),"format":cell(row,cf,12),
                                     "folder":cell(row,cfo,40),"status":st,"kind":kind})
        cov["doc_stats"]={"total":len(cov["documents"]),"vision":v,"ocr":t,"read":r}
    # X01 — OCR rysunków (vision)
    rows=_sheet_rows(wb,"X01_OCR","OCR_PROJEKT","OCR_RYSUN")
    if rows:
        h1=" ".join(str(c) for c in rows[0] if c) if rows else ""
        m=re.search(r"vision[^),]*",h1,re.I)
        if m: cov["model"]=m.group(0).strip()
        hi=_hdr_idx(rows,"nr rys","rys")
        items=[]
        if hi is not None:
            head=[str(x).strip().lower() if x else "" for x in rows[hi]]
            cn=_col(head,"nr rys","rys",default=0); cty=_col(head,"typ",default=1); cu=_col(head,"uwagi")
            for row in rows[hi+1:]:
                nr=cell(row,cn,24)
                if not nr: continue
                items.append({"nr":nr,"typ":cell(row,cty,14),"uwagi":cell(row,cu,60)})
        cov["ocr_rysunki"]={"count":len(items),"items":items[:50]}
    # X03 — OCR decyzji / skanów
    rows=_sheet_rows(wb,"X03_OCR","OCR_DECYZJE","DECYZJE_SKAN")
    if rows:
        hi=_hdr_idx(rows,"dokument"); items=[]
        if hi is not None:
            head=[str(x).strip().lower() if x else "" for x in rows[hi]]
            cd=_col(head,"dokument",default=0); csg=_col(head,"sygnatura",default=1); cst=_col(head,"status")
            for row in rows[hi+1:]:
                dk=cell(row,cd,60)
                if not dk: continue
                items.append({"dokument":dk,"sygnatura":cell(row,csg,30),"status":cell(row,cst,14)})
        cov["ocr_decyzje"]={"count":len(items),"items":items[:30]}
    # X02 — kontrola krzyżowa projekt vs przedmiar
    rows=_sheet_rows(wb,"X02_CROSS","CROSS_REF")
    if rows:
        hi=_hdr_idx(rows,"poz. przedm","poz.przedm","przedm")
        ok=diff=0; items=[]
        if hi is not None:
            head=[str(x).strip().lower() if x else "" for x in rows[hi]]
            cst=_col(head,"status"); co=_col(head,"opis",default=1)
            for row in rows[hi+1:]:
                if not any(row): continue
                st=cell(row,cst,18).upper()
                if not st: continue
                if "OK" in st: ok+=1
                else: diff+=1
                items.append({"opis":cell(row,co,60),"status":st})
        cov["crossref"]={"count":ok+diff,"ok":ok,"diff":diff,"items":items[:50]}
    # 07 — rozbieżności
    rows=_sheet_rows(wb,"07_ROZBIEZ","ROZBIEZN")
    hi=_hdr_idx(rows,"element") if rows else None
    if hi is not None:
        head=[str(x).strip().lower() if x else "" for x in rows[hi]]
        ci=_col(head,"id",default=0); ce=_col(head,"element",default=1); cst=_col(head,"status"); cop=_col(head,"opis")
        for row in rows[hi+1:]:
            el=cell(row,ce,50)
            if not el: continue
            cov["rozbieznosci"].append({"id":cell(row,ci,8),"element":el,
                "status":cell(row,cst,16),"opis":cell(row,cop,110)})
        cov["rozbieznosci"]=cov["rozbieznosci"][:40]
    # 08 — braki przedmiaru
    rows=_sheet_rows(wb,"08_BRAKI","BRAKI_PRZEDMIAR")
    hi=_hdr_idx(rows,"pozycja") if rows else None
    if hi is not None:
        head=[str(x).strip().lower() if x else "" for x in rows[hi]]
        ci=_col(head,"id",default=0); cpz=_col(head,"pozycja",default=1); cw=_col(head,"wartość","wartosc"); cu=_col(head,"uwaga")
        for row in rows[hi+1:]:
            pz=cell(row,cpz,60)
            if not pz: continue
            cov["braki"].append({"id":cell(row,ci,8),"pozycja":pz,
                "wartosc":num(row[cw]) if cw is not None and len(row)>cw else None,"uwaga":cell(row,cu,90)})
        cov["braki"]=cov["braki"][:40]
    # E01-E10 — eksperci branżowi
    EXP={"E01":"Prawnik zamówień (Pzp)","E02":"Radca umowy","E03":"Konstruktor","E04":"Architekt-konserwator",
         "E05":"Kosztorysant","E06":"Zakupowiec","E07":"Elektryk","E08":"Sanitarny","E09":"Wentylacja","E10":"Stolarka konserw."}
    for s in wb.sheetnames:
        for k,name in EXP.items():
            if s.upper().startswith(k) and name not in cov["experts"]: cov["experts"].append(name)
    return cov
